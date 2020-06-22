import psycopg2
import openpyxl


def conectar():
    """
    Função para conectar ao servidor
    """
    try:
        conn = psycopg2.connect(
            database='ppostgresql',
            host='localhost',
            user='geek',
            password='university'
        )
        return conn
    except psycopg2.Error as e:
        print(f'Erro na conexão ao PostgreSQL Server: {e}.')


def desconectar(conn):
    """ 
    Função para desconectar do servidor.
    """
    if conn:
        conn.close()


def listar():
    """
    Função para listar os produtos
    """
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM produtos')
    produtos = cursor.fetchall()
    if len(produtos) > 0:
        print('Listando produtos ...')
        print('-' * 30)
        for produto in produtos:
            print(f'ID: {produto[0]}')
            print(f'Produto: {produto[1]}')
            print(f'Preço: {produto[2]}')
            print(f'Estoque: {produto[3]}')
            print('-' * 30)
    else:
        print('Não existem produtos cadastrados!')
    desconectar(conn)


def inserir():
    """
    Função para inserir um produto
    """
    conn = conectar()
    cursor = conn.cursor()

    nome = str(input('Informe o nome do produto: '))
    preco = float(input('Informe o preço do produto: '))
    estoque = int(input('Informe a quantidade em estoque: '))

    cursor.execute(f"INSERT INTO produtos (nome, preco, estoque) VALUES ('{nome}', {preco}, {estoque})")
    conn.commit()

    if cursor.rowcount == 1:
        print(f'O produto {nome} foi inserido com sucesso!')
    else:
        print('Não foi possível inserir o produto!')
    desconectar(conn)


def atualizar():
    """
    Função para atualizar um produto
    """
    conn = conectar()
    cursor = conn.cursor()

    codigo = int(input('Informe o código do produto: '))
    nome = input('Informe o novo nome do produto: ')
    preco = float(input('Informe o novo preço do produto: '))
    estoque = int(input('Informe a nova quantidade em estoque: '))

    cursor.execute(f"UPDATE produtos SET nome='{nome}', preco={preco}, estoque={estoque} WHERE id={codigo}")
    conn.commit()

    if cursor.rowcount == 1:
        print(f'O produto {nome} foi atualizado com sucesso!')
    else:
        print('Erro ao atualizar o produto!')
    desconectar(conn)


def deletar():
    """
    Função para deletar um produto
    """
    conn = conectar()
    cursor = conn.cursor()

    codigo = int(input('Informe o código do produto: '))

    cursor.execute(f'DELETE FROM produtos WHERE id={codigo}')
    conn.commit()

    if cursor.rowcount == 1:
        print('Produto excluido com sucesso!')
    else:
        print(f'Erro ao excluir o produto com ID {codigo}.')
    desconectar(conn)


def escrever():
    conn = conectar()
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM produtos')
    produtos = cursor.fetchall()

    for produto in produtos:
        with open('teste2.txt', 'a') as arq:
            arq.write(f'{produto[0]}')
            arq.write(f'\t{produto[1]}')
            arq.write(f'{" " * (25 - len(str(produto[1])))}{produto[2]}')
            arq.write(f'{" " * (25 - len(str(produto[2])))}{produto[3]}\n')
    desconectar(conn)


def excel():
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM produtos')
    produtos = cursor.fetchall()

    planilia = openpyxl.Workbook()
    planilia.create_sheet('Página1', 0)
    planilia1 = planilia['Página1']

    for c in range(1, len(produtos) + 1):
        for i in range(1, 5):
            planilia1.cell(c, i).value = produtos[c - 1][i - 1]
    planilia.save('teste.xlsx')


def menu():
    """
    Função para gerar o menu inicial
    """
    print('=========Gerenciamento de Produtos==============')
    print('Selecione uma opção: ')
    print('1 - Listar produtos.')
    print('2 - Inserir produtos.')
    print('3 - Atualizar produto.')
    print('4 - Deletar produto.')
    print('5 - Escrever em arquivo de texto.')
    print('6 - Escrever em arquivo excel.')
    opcao = int(input())
    if opcao in [1, 2, 3, 4, 5, 6]:
        if opcao == 1:
            listar()
        elif opcao == 2:
            inserir()
        elif opcao == 3:
            atualizar()
        elif opcao == 4:
            deletar()
        elif opcao == 5:
            escrever()
        elif opcao == 6:
            excel()
        else:
            print('Opção inválida')
    else:
        print('Opção inválida')
